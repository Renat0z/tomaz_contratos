import os
import fitz  # PyMuPDF
import pandas as pd
import re
from io import StringIO
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# Configuração da chave da API OpenAI
openai_api_key = 'sk-proj-IjBLQbETogwa4x92RJ68T3BlbkFJmjzCOYbHNypxos121z7U'

# Criar a instância do modelo
llm = ChatOpenAI(
    model="gpt-4o-mini"
)

def chat_with_openai(prompt):
    messages = [HumanMessage(content=prompt)]
    response = llm.invoke(messages)
    return response.content

def extract_json(response):
    json_match = re.search(r'\{.*?\}', response, re.DOTALL)
    if json_match:
        return json_match.group(0)
    else:
        return None

def read_pdfs_from_directory(directory):
    pdf_files = [f for f in os.listdir(directory) if f.endswith('.pdf')]
    print("PDFs encontrados:", pdf_files)
    return pdf_files

def get_pdfs_from_excel(file_path):
    try:
        if os.path.exists(file_path):
            df = pd.read_excel(file_path, sheet_name='contratos', engine='openpyxl')
            if 'Contrato' in df.columns and 'Link do Contrato' in df.columns:
                return df['Contrato'].tolist(), df
        else:
            df = pd.DataFrame(columns=['Contrato', 'Link do Contrato'])
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='contratos', index=False)
            return [], df
    except (FileNotFoundError, InvalidFileException, BadZipFile) as e:
        print(f"Erro ao ler o arquivo Excel: {e}")
        return [], pd.DataFrame(columns=['Contrato', 'Link do Contrato'])

def compare_pdfs(directory, excel_file_path):
    directory_pdfs = read_pdfs_from_directory(directory)
    excel_pdfs, _ = get_pdfs_from_excel(excel_file_path)
    not_in_excel = [pdf for pdf in directory_pdfs if pdf not in excel_pdfs]
    print("PDFs no diretório que não estão no Excel:", not_in_excel)
    return not_in_excel

def extract_text_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    return text

def save_text_to_file(text, txt_path):
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write(text)
    print(f"Texto salvo em: {txt_path}")

def log_message(message, log_file='process.log'):
    with open(log_file, 'a', encoding='utf-8') as log:
        log.write(message + '\n')

def process_text_with_openai(text):
    user_input = f"""Extraia os dados do seguinte texto e retorne em formato JSON:\n
    {text}
    Baseando-se nas informações extraídas do contrato, aqui está a definição das chaves no JSON, seguida de uma explicação para encontrar as exatas informações nos contratos:
    exemplo:
      'EQL': 'Unidade',
      'Dt Contrato': 'Data do contrato',
      'CPF/CNPJ': 'CPF/CNPJ do comprador',
      'Data_Vencimento': 'Data de vencimento da parcela',
      'Data_Pagamento': 'Data de pagamento da parcela',
      'Referência': 'Descrição da parcela',
      'Nome': 'Nome do comprador',
      'Valor': 'Valor da parcela',
      'Cm+Juros+Multa': 'Correção monetária, juros e multa',
      'D.Banco': 'Banco destinatário',
      'Total': 'Total pago'
    Explicação para Encontrar Informações no Contrato
    EQL (Unidade):

    Localização: Seção que menciona a unidade do imóvel no contrato.
    Exemplo de Frase: 'Unidade: F3'
    Dt Contrato (Data do contrato):

    Localização: Primeiras páginas, geralmente no Quadro-Resumo.
    Exemplo de Frase: 'Data do contrato: 01 de setembro de 2023'
    CPF/CNPJ (CPF/CNPJ do comprador):

    Localização: Seção que descreve os dados do comprador.
    Exemplo de Frase: 'CPF/MF: 215.023.498-09'
    Data_Vencimento (Data de vencimento da parcela):

    Localização: Seção de Condições da Venda e Forma de Pagamento.
    Exemplo de Frase: 'vencimento em 15/02/2024'
    Data_Pagamento (Data de pagamento da parcela):

    Localização: Pode variar, geralmente próximo à informação de vencimento.
    Exemplo de Frase: 'Data de pagamento: 09/02/2024'
    Referência (Descrição da parcela):

    Localização: Seção de Condições da Venda e Forma de Pagamento ou especificações da parcela.
    Exemplo de Frase: 'Prestações mensais'
    Nome (Nome do comprador):

    Localização: Seção de Dados do Comprador.
    Exemplo de Frase: 'Nome comprador: JAMES ANDERSON MORAES'
    Valor (Valor da parcela):

    Localização: Seção de Condições da Venda e Forma de Pagamento.
    Exemplo de Frase: 'Valor da parcela: R$ 1.977,38'
    Cm+Juros+Multa (Correção monetária, juros e multa):

    Localização: Seção de Condições da Venda e Forma de Pagamento, detalhes das parcelas.
    Exemplo de Frase: 'Incidência de juros de 0,95% a.m.'
    D.Banco (Banco destinatário):

    Localização: Pode não estar explicitamente listado no contrato, verificar boletos ou instruções de pagamento.
    Exemplo de Frase: 'Pagamento através de boleto bancário'
    Total (Total pago):

    Localização: Geralmente aparece no resumo de pagamentos ou totais das parcelas.
    Exemplo de Frase: 'Total: R$ 2.008,61'
    Utilizando essa estrutura e as frases exemplo, uma LLM pode ser treinada para identificar e extrair essas informações de forma consistente nos contratos.

    Coloque apenas um valor em cada chave json. Esses valores serão adicionados em um xlsx e por isso pode haver apenas um valor por linha.
    
    Siga o EXATO modelo abaixo:
      'EQL': ' 16.  F.   3',
      'Dt Contrato': '9/1/2023',
      'CPF/CNPJ': '215023498-09',
      'Data_Vencimento': '2/15/2024',
      'Data_Pagamento': '2/9/2024',
      'Referência': 'Prestação II 1 / 56',
      'Nome': 'JAMES ANDERSON MORAES',
      'Valor': '  2,008.61 ',
      'Cm+Juros+Multa': '  -   ',
      'D.Banco': '  -   ',
      'Total': ' R$  2,008.61 '
    """
    response = chat_with_openai(user_input)
    return extract_json(response)

def save_json_to_excel(json_content, excel_file_path):
    try:
        data = pd.read_json(StringIO(json_content), typ='series').to_frame().T
    except ValueError as e:
        print(f"Erro ao converter JSON para DataFrame: {e}")
        return

    try:
        # Abre o arquivo Excel existente, preservando os dados existentes
        book = openpyxl.load_workbook(excel_file_path)
        if 'recebiveis' in book.sheetnames:
            df_recebiveis = pd.read_excel(excel_file_path, sheet_name='recebiveis', engine='openpyxl')
        else:
            # Cria um novo DataFrame se a planilha não existir
            df_recebiveis = pd.DataFrame(columns=['EQL', 'Dt Contrato', 'CPF/CNPJ', 'Data_Vencimento', 'Data_Pagamento', 'Referência', 'Nome', 'Valor', 'Cm+Juros+Multa', 'D.Banco', 'Total'])
            
        # Concatena os dados JSON ao DataFrame existente ou ao novo DataFrame
        df_recebiveis = pd.concat([df_recebiveis, data], ignore_index=True)
            
        # Ordena o DataFrame pela coluna 'EQL' por ordem alfabética
        df_recebiveis = df_recebiveis.sort_values(by='EQL', ascending=True)
            
        # Salva o DataFrame atualizado de volta ao Excel na planilha 'recebiveis'
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_recebiveis.to_excel(writer, sheet_name='recebiveis', index=False)
        log_message("Dados do JSON adicionados ao Excel com sucesso.")
        print("Dados do JSON adicionados ao Excel com sucesso.")
    except (FileNotFoundError, InvalidFileException, BadZipFile) as e:
        # Se o arquivo Excel não existir ou estiver corrompido, cria um novo com os dados JSON
        data.to_excel(excel_file_path, sheet_name='recebiveis', index=False, engine='openpyxl')
        log_message(f"Arquivo Excel criado e dados do JSON adicionados com sucesso. Erro encontrado: {e}")
        print(f"Arquivo Excel criado e dados do JSON adicionados com sucesso. Erro encontrado: {e}")

def update_excel_with_pdf_info(pdf_file, excel_file_path, df):
    link = os.path.abspath(pdf_file)
    new_row = pd.DataFrame({'Contrato': [os.path.basename(pdf_file)], 'Link do Contrato': [link]})
    df = pd.concat([df, new_row], ignore_index=True)
    try:
        # Abre o arquivo Excel existente, preservando os dados existentes
        book = openpyxl.load_workbook(excel_file_path)
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Preserva os dados existentes em todas as planilhas
            for sheet in book.sheetnames:
                existing_df = pd.read_excel(excel_file_path, sheet_name=sheet, engine='openpyxl')
                existing_df.to_excel(writer, sheet_name=sheet, index=False)
            # Atualiza a planilha 'contratos' com os novos dados
            df.to_excel(writer, sheet_name='contratos', index=False)
    except (FileNotFoundError, InvalidFileException, BadZipFile) as e:
        # Se o arquivo Excel não existir, cria um novo com a planilha 'contratos'
        df.to_excel(excel_file_path, sheet_name='contratos', index=False, engine='openpyxl')
        log_message(f"Erro ao abrir arquivo Excel existente. Criando novo arquivo. Erro encontrado: {e}")
    return df

def process_pdfs(pdf_files, directory, output_directory, excel_file_path):
    existing_pdfs, df = get_pdfs_from_excel(excel_file_path)

    for pdf_file in pdf_files:
        pdf_path = os.path.join(directory, pdf_file)
        txt_file = pdf_file.replace('.pdf', '.txt')
        txt_path = os.path.join(output_directory, txt_file)

        text = extract_text_from_pdf(pdf_path)
        save_text_to_file(text, txt_path)

        json_content = process_text_with_openai(text)
        if json_content:
            save_json_to_excel(json_content, excel_file_path)
            df = update_excel_with_pdf_info(pdf_path, excel_file_path, df)
            log_message(f"Arquivo PDF {pdf_file} processado e informações adicionadas ao Excel.")
        else:
            log_message(f"Falha ao extrair JSON do texto do arquivo PDF {pdf_file}.")
        print(f"Arquivo PDF {pdf_file} processado e informações adicionadas ao Excel.")

# Exemplo de uso
directory_path = 'contratos'
output_directory_path = 'contratos/textos'
excel_file_path = 'clientes.xlsx'

pdfs_to_extract = compare_pdfs(directory_path, excel_file_path)
process_pdfs(pdfs_to_extract, directory_path, output_directory_path, excel_file_path)

# Carregar o arquivo Excel modificado para verificar os resultados
try:
    if 'recebiveis' in pd.ExcelFile(excel_file_path).sheet_names:
        modified_df = pd.read_excel(excel_file_path, sheet_name='recebiveis')
        print(modified_df.head())
    else:
        print("Worksheet 'recebiveis' não foi encontrado.")
except (FileNotFoundError, InvalidFileException, BadZipFile) as e:
    print(f"Erro ao abrir arquivo Excel: {e}")
